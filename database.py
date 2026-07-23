import openpyxl
from openpyxl import load_workbook
from collections import Counter
import os
from datetime import datetime

EXCEL_FILE = "wingo_30s_data.xlsx"

def init_excel():
    """Excel ফাইল তৈরি করবে (যদি না থাকে) সব শীটসহ"""
    if os.path.exists(EXCEL_FILE):
        return
    wb = openpyxl.Workbook()
    
    # All Rounds
    ws1 = wb.active
    ws1.title = "All Rounds"
    ws1.append(["Period", "Number", "Size", "Prediction", "Result", "Range_Pred"])
    
    # Summary
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Statistic", "Value"])
    ws2.append(["Total Rounds (All Time)", 0])
    ws2.append(["Total Patterns (all lengths)", 0])
    ws2.append(["Unique Lengths", 0])
    ws2.append(["Last Updated", ""])
    ws2.append([])
    ws2.append(["Length", "Count", "Most Frequent Pattern", "Frequency"])
    for length in range(2, 7):
        ws2.append([length, 0, "", 0])
    
    # Length 2 to 6
    for length in range(2, 7):
        ws = wb.create_sheet(f"Length {length}")
        ws.append(["Pattern", "Frequency"])
    
    wb.save(EXCEL_FILE)

def save_round(period, number, size, prediction, result, range_pred):
    """প্রতি রাউন্ডের ডেটা Excel-এ যোগ করবে ও সব শীট আপডেট করবে"""
    wb = load_workbook(EXCEL_FILE)
    ws_all = wb["All Rounds"]
    ws_all.append([period, number, size, prediction, result, range_pred])
    
    # সব ডেটা সংগ্রহ
    all_data = []
    for row in ws_all.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[2] in ("BIG", "SMALL"):
            all_data.append(row[2])
    
    # প্যাটার্ন কাউন্ট
    length_patterns = {}
    for length in range(2, 7):
        patterns = []
        for i in range(len(all_data) - length + 1):
            patterns.append("-".join(all_data[i:i+length]))
        length_patterns[length] = Counter(patterns)
    
    # Summary আপডেট
    ws_sum = wb["Summary"]
    ws_sum["B2"] = len(all_data)
    total_patterns = sum(sum(c.values()) for c in length_patterns.values())
    ws_sum["B3"] = total_patterns
    ws_sum["B4"] = sum(1 for c in length_patterns.values() if c)
    ws_sum["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    row_idx = 8
    for length in range(2, 7):
        counter = length_patterns[length]
        if counter:
            most = counter.most_common(1)[0]
            ws_sum.cell(row=row_idx, column=1, value=length)
            ws_sum.cell(row=row_idx, column=2, value=len(counter))
            ws_sum.cell(row=row_idx, column=3, value=most[0])
            ws_sum.cell(row=row_idx, column=4, value=most[1])
        else:
            ws_sum.cell(row=row_idx, column=1, value=length)
            ws_sum.cell(row=row_idx, column=2, value=0)
            ws_sum.cell(row=row_idx, column=3, value="")
            ws_sum.cell(row=row_idx, column=4, value=0)
        row_idx += 1
    
    # Length শীট আপডেট
    for length in range(2, 7):
        ws_len = wb[f"Length {length}"]
        ws_len.delete_rows(2, ws_len.max_row)
        for pattern, freq in length_patterns[length].most_common():
            ws_len.append([pattern, freq])
    
    wb.save(EXCEL_FILE)

def load_recent_history(limit=None):
    """All Rounds থেকে সব রেকর্ড লোড করবে (limit প্যারামিটার ইগনোর)"""
    if not os.path.exists(EXCEL_FILE):
        return []
    wb = load_workbook(EXCEL_FILE)
    ws = wb["All Rounds"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    data = []
    for row in rows:
        if len(row) < 6:
            row = list(row) + [None] * (6 - len(row))
        data.append(tuple(row))
    return data